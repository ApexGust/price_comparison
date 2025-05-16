import datetime
import json
import sys

import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def load_and_prepare_data(file_path, supplier_name_from_ui, product_name_col, spec_col_name, price_col):  # spec_col_name can be None
    try:
        if not os.path.exists(file_path):
            messagebox.showerror("文件错误", f"文件未找到: {file_path}")
            return None
        df = pd.read_excel(file_path)

        if product_name_col in df.columns:
            df[product_name_col] = df[product_name_col].ffill() # Forward fill 品名列
        # 如果规格列也可能存在合并单元格，同样处理 spec_col_name (如果提供了)
        if spec_col_name and spec_col_name.strip() and spec_col_name in df.columns:
             df[spec_col_name] = df[spec_col_name].ffill() # Forward fill 规格列

        cols_to_read = [product_name_col, price_col]
        rename_map = {product_name_col: '品名', price_col: '价格'}

        if product_name_col not in df.columns:
            messagebox.showerror("列名错误", f"在 {supplier_name_from_ui} 的文件 '{os.path.basename(file_path)}' 中未找到品名列: '{product_name_col}'.")
            return None
        if price_col not in df.columns:
            messagebox.showerror("列名错误", f"在 {supplier_name_from_ui} 的文件 '{os.path.basename(file_path)}' 中未找到价格列: '{price_col}'.")
            return None

        has_spec_col = False
        if spec_col_name and spec_col_name.strip():  # 如果提供了非空的规格列名
            if spec_col_name not in df.columns:
                messagebox.showerror("列名错误", f"在 {supplier_name_from_ui} 的文件 '{os.path.basename(file_path)}' 中未找到指定的规格列: '{spec_col_name}'.")
                return None
            cols_to_read.insert(1, spec_col_name)  # 在品名和价格之间插入规格列
            rename_map[spec_col_name] = '规格'
            has_spec_col = True

        df_selected = df[cols_to_read].copy()
        df_selected.rename(columns=rename_map, inplace=True)

        df_selected['供应商'] = supplier_name_from_ui
        df_selected['品名'] = df_selected['品名'].astype(str).str.strip()
        if has_spec_col:
            df_selected['规格'] = df_selected['规格'].astype(str).str.strip()
        else:
            df_selected['规格'] = ""  # 如果没有规格列，则规格默认为空字符串

        df_selected['价格'] = pd.to_numeric(df_selected['价格'], errors='coerce')

        df_selected['产品标识符'] = df_selected.apply(
            lambda row: f"{row['品名']}|{row['规格']}" if row['规格'] else row['品名'],
            axis=1
        )

        df_selected.dropna(subset=['品名', '价格', '产品标识符'], inplace=True)
        df_selected = df_selected[df_selected['价格'] > 0]

        if df_selected.empty:
            messagebox.showwarning("数据警告", f"{supplier_name_from_ui} 的文件 '{os.path.basename(file_path)}' 中没有找到有效的带价格的产品数据。")
        return df_selected
    except Exception as e:
        messagebox.showerror("加载错误", f"加载 {supplier_name_from_ui} 的文件 '{os.path.basename(file_path)}' 失败: {e}")
        return None


def parse_procurement_input(procurement_text):
    # (这个函数在上一版中已确认无误，保持不变)
    procurement_dict_display = {}
    procurement_dict_internal = {}
    lines = procurement_text.strip().split('\n')
    if not lines or (len(lines) == 1 and not lines[0].strip()):
        messagebox.showerror("输入错误", "采购清单不能为空。")
        return None, None
    example_text = "例如:\n土豆,70cm,100\n苹果,大,50\n香蕉,小,20\n白菜,,30 (如无规格则第二项留空)"
    if lines[0].strip().lower().startswith("例如:"):
        messagebox.showinfo("提示", f"请注意：采购清单格式为：品名,规格,数量\n{example_text}")
        if len(lines) <= 1 or (len(lines) > 1 and not lines[1].strip()):
            messagebox.showerror("输入错误", "采购清单不能为空，示例行之后需要有实际采购条目。")
            return None, None
        lines = lines[1:]
    valid_entry_found = False
    for i, line in enumerate(lines):
        line = line.strip()
        if not line: continue
        parts = line.split(',')
        if len(parts) != 3:
            messagebox.showerror("输入错误", f"采购清单第 {i + 1} 行 (内容: '{line}') 格式错误。\n请使用 '品名,规格,数量' 格式。\n{example_text}")
            return None, None
        product_name = parts[0].strip()
        spec_name_input = parts[1].strip()
        quantity_str = parts[2].strip()
        if not product_name:
            messagebox.showerror("输入错误", f"采购清单第 {i + 1} 行 (内容: '{line}') 的品名不能为空。")
            return None, None
        try:
            quantity = int(quantity_str)
            if quantity <= 0:
                messagebox.showerror("输入错误", f"采购清单第 {i + 1} 行产品 '{product_name}' 的数量 ('{quantity_str}') 必须为正整数。")
                return None, None
            internal_key = f"{product_name}|{spec_name_input}" if spec_name_input else product_name
            display_key = f"{product_name} ({spec_name_input})" if spec_name_input else product_name
            if internal_key in procurement_dict_internal:
                procurement_dict_internal[internal_key]['数量'] += quantity
            else:
                procurement_dict_internal[internal_key] = {'品名': product_name, '规格': spec_name_input, '数量': quantity}
            if display_key in procurement_dict_display:
                procurement_dict_display[display_key] += quantity
            else:
                procurement_dict_display[display_key] = quantity
            valid_entry_found = True
        except ValueError:
            messagebox.showerror("输入错误", f"采购清单第 {i + 1} 行产品 '{product_name}' 的数量 ('{quantity_str}') 格式错误。数量应为整数。")
            return None, None
    if not valid_entry_found:
        messagebox.showerror("输入错误", "采购清单中未找到有效的采购条目，或所有条目格式均不正确。")
        return None, None
    return procurement_dict_internal, procurement_dict_display


def generate_purchase_plan(supplier_dataframes_dict, procurement_needs_internal, current_supplier_display_names):
    # (这个函数在上一版中已确认无误，保持不变)
    valid_supplier_dfs_list = [df for df in supplier_dataframes_dict.values() if df is not None and not df.empty]
    if not valid_supplier_dfs_list:
        messagebox.showerror("数据错误", "没有可用的供应商数据进行比价。")
        return pd.DataFrame(), {}, ["没有加载到任何有效的供应商报价数据。"]
    all_prices_df = pd.concat(valid_supplier_dfs_list, ignore_index=True)
    if all_prices_df.empty:
        messagebox.showerror("数据错误", "所有供应商的报价数据均为空或无效。")
        return pd.DataFrame(), {}, ["所有供应商的报价数据均为空或无效。"]
    purchase_details_list = []
    supplier_totals = {name: 0.0 for name in current_supplier_display_names if name in all_prices_df['供应商'].unique()}
    not_found_in_any_supplier = []
    for internal_key, need_details in procurement_needs_internal.items():
        # print(f"调试：采购需求 internal_key: '{internal_key}'")
        product_name_needed = need_details['品名']
        # relevant_offers = all_prices_df[all_prices_df['品名'].str.strip().str.lower() == product_name_needed.strip().lower()]
        # if not relevant_offers.empty:
        #     print(f"调试：Excel中找到的 '{product_name_needed}' 的产品标识符: \n{relevant_offers[['供应商', '品名', '规格', '产品标识符']]}")
        # else:
        #     print(f"调试：Excel中未找到品名 '{product_name_needed}'")
        # product_offers_for_this_item = all_prices_df[
        #     all_prices_df['产品标识符'].str.strip().str.lower() == internal_key.strip().lower()
        #     ]
        # if product_offers_for_this_item.empty:
        #     print(f"调试：使用 internal_key '{internal_key}' 未在 all_prices_df 中匹配到任何产品。")
        spec_needed = need_details['规格']
        quantity_needed = need_details['数量']
        display_name_for_output = f"{product_name_needed} ({spec_needed})" if spec_needed else product_name_needed
        product_offers_for_this_item = all_prices_df[
            all_prices_df['产品标识符'].str.strip().str.lower() == internal_key.strip().lower()
            ]
        if product_offers_for_this_item.empty:
            not_found_in_any_supplier.append(display_name_for_output)
            continue
        best_offer_series = product_offers_for_this_item.loc[product_offers_for_this_item['价格'].idxmin()]
        min_price = best_offer_series['价格']
        chosen_supplier = best_offer_series['供应商']
        sub_total = min_price * quantity_needed
        price_comparison_details = {}
        for sup_name in current_supplier_display_names:
            offer_from_sup = product_offers_for_this_item[product_offers_for_this_item['供应商'] == sup_name]
            if not offer_from_sup.empty:
                price_comparison_details[sup_name] = offer_from_sup['价格'].iloc[0]
            else:
                price_comparison_details[sup_name] = "未报价"
        purchase_details_list.append({
            '产品显示名称': display_name_for_output,
            '品名': product_name_needed,
            '规格': spec_needed,
            '采购数量': quantity_needed,
            '选择的供应商': chosen_supplier,
            '单价': min_price,
            '金额': sub_total,
            '比价详情': price_comparison_details
        })
        if chosen_supplier in supplier_totals:
            supplier_totals[chosen_supplier] += sub_total
    purchase_df = pd.DataFrame(purchase_details_list)
    notes = []
    if not_found_in_any_supplier:
        notes.append(f"以下产品在所有供应商报价中均未找到: {', '.join(not_found_in_any_supplier)}.")
    supplier_totals = {k: v for k, v in supplier_totals.items() if v > 0 or (not purchase_df.empty and k in purchase_df['选择的供应商'].unique())}
    return purchase_df, supplier_totals, notes


class ProcurementApp:
    MIN_SUPPLIERS = 2
    MAX_SUPPLIERS = 5
    INITIAL_SUPPLIERS = 2
    CACHE_FILE_NAME = "procurement_list_cache.json"

    def __init__(self, root):
        self.root = root
        self.root.title(f"智能采购比价系统 V2.6")
        self.root.geometry("1150x750")  # 调整高度，因为底部汇总区移除了

        self.current_purchase_df = pd.DataFrame()
        self.DEFAULT_PRODUCT_NAME_COL = '品名'
        self.DEFAULT_SPEC_COL = '规格'
        self.DEFAULT_PRICE_COL = '价格'

        self.supplier_entries = []
        self.supplier_frame_widgets = {}

        # 用于导出Excel时获取上次运行的汇总数据
        self.last_run_supplier_totals_dict = {}
        self.last_run_notes_list = []
        self.last_run_grand_total_cost = 0.0

        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(expand=True, fill=tk.BOTH)

        self.step1_frame = ttk.LabelFrame(main_frame, text="步骤 1: 选择供应商报价文件并指定列名", padding="10")
        self.step1_frame.pack(fill=tk.X, pady=5)

        self.suppliers_dynamic_area = ttk.Frame(self.step1_frame)
        self.suppliers_dynamic_area.pack(fill=tk.X)

        supplier_buttons_frame = ttk.Frame(self.step1_frame)
        supplier_buttons_frame.pack(fill=tk.X, pady=5)
        self.add_supplier_button = ttk.Button(supplier_buttons_frame, text="✚ 添加供应商", command=self.add_supplier_input)
        self.add_supplier_button.pack(side=tk.LEFT, padx=5)
        self.remove_supplier_button = ttk.Button(supplier_buttons_frame, text="➖ 移除最后一个供应商", command=self.remove_last_supplier_input)
        self.remove_supplier_button.pack(side=tk.LEFT, padx=5)

        for _ in range(self.INITIAL_SUPPLIERS):
            self._add_supplier_row_ui()
        self._update_add_remove_buttons_state()

        ttk.Separator(self.step1_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10, expand=True)

        excel_cols_frame = ttk.Frame(self.step1_frame)
        excel_cols_frame.pack(fill=tk.X, pady=5)
        ttk.Label(excel_cols_frame, text="Excel品名列名:").grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)
        self.product_name_col_var = tk.StringVar(value=self.DEFAULT_PRODUCT_NAME_COL)
        ttk.Entry(excel_cols_frame, textvariable=self.product_name_col_var, width=20).grid(row=0, column=1, padx=5, pady=2, sticky=tk.W)
        ttk.Label(excel_cols_frame, text="Excel规格列名:").grid(row=1, column=0, padx=5, pady=2, sticky=tk.W)
        self.spec_col_var = tk.StringVar(value=self.DEFAULT_SPEC_COL)
        ttk.Entry(excel_cols_frame, textvariable=self.spec_col_var, width=20).grid(row=1, column=1, padx=5, pady=2, sticky=tk.W)
        ttk.Label(excel_cols_frame, text="Excel价格列名:").grid(row=2, column=0, padx=5, pady=2, sticky=tk.W)
        self.price_col_var = tk.StringVar(value=self.DEFAULT_PRICE_COL)
        ttk.Entry(excel_cols_frame, textvariable=self.price_col_var, width=20).grid(row=2, column=1, padx=5, pady=2, sticky=tk.W)

        step2_frame = ttk.LabelFrame(main_frame, text="步骤 2: 输入采购需求", padding="10")
        step2_frame.pack(fill=tk.X, pady=5)
        ttk.Label(step2_frame, text="输入采购清单 (每行格式: 品名,规格,数量):").pack(anchor=tk.W)
        self.procurement_needs_text = tk.Text(step2_frame, height=8, width=70, relief=tk.SOLID, borderwidth=1)
        self.procurement_needs_text.pack(fill=tk.X, pady=5)
        self.procurement_needs_text.insert(tk.END, "例如:\n土豆,70cm,100\n苹果,大,50\n香蕉,小,20\n白菜,,30 (如无规格则第二项留空)")

        action_frame = ttk.Frame(main_frame, padding="5")
        action_frame.pack(fill=tk.X, pady=5)
        self.run_button = ttk.Button(action_frame, text="开始比价并生成采购单", command=self.run_analysis, style="Accent.TButton")
        self.run_button.pack(side=tk.LEFT, padx=5)
        self.export_button = ttk.Button(action_frame, text="导出采购单到Excel", command=self.export_to_excel, state=tk.DISABLED)
        self.export_button.pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="清空输入", command=self.clear_inputs).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="退出", command=root.quit).pack(side=tk.RIGHT, padx=5)

        results_frame = ttk.LabelFrame(main_frame, text="结果展示", padding="10")
        results_frame.pack(expand=True, fill=tk.BOTH, pady=5)
        ttk.Label(results_frame, text="采购详情 (按供应商):", font="-weight bold").pack(anchor=tk.W)

        self.purchase_table_container = ttk.Frame(results_frame)
        self.purchase_table_container.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        self.purchase_table = None

        # 底部汇总信息区域已移除

        style = ttk.Style()
        style.configure("Accent.TButton", foreground="white", background="green", font=("-weight bold"))

        self._load_cached_procurement_list()

        # --- 修改点：设置窗口关闭时的回调 ---
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def _get_cache_file_path(self):
        """获取缓存文件的绝对路径 (与exe同目录)"""
        # 对于 PyInstaller 打包的 exe，sys.executable 是 exe 的路径
        # 对于直接运行的 .py 文件，os.path.dirname(__file__) 是脚本所在目录
        try:
            # 如果是 PyInstaller 打包的单文件 exe，sys._MEIPASS 是临时解压目录
            # 我们希望缓存文件与 exe 在同一目录，而不是临时目录
            if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                # For PyInstaller one-file bundle, executable is in sys.executable
                base_path = os.path.dirname(sys.executable)
            else:
                # For .py script or PyInstaller one-folder bundle
                base_path = os.path.dirname(os.path.abspath(__file__))
        except NameError: # __file__ might not be defined in some contexts (e.g. interactive)
             base_path = os.getcwd() # Fallback to current working directory
        return os.path.join(base_path, self.CACHE_FILE_NAME)

    def _load_cached_procurement_list(self):
        cache_path = self._get_cache_file_path()
        try:
            if os.path.exists(cache_path):
                with open(cache_path, 'r', encoding='utf-8') as f:
                    cached_data = json.load(f)
                    procurement_text = cached_data.get("procurement_list", "")
                    if procurement_text: # 确保不是空字符串再覆盖
                        self.procurement_needs_text.delete("1.0", tk.END)
                        self.procurement_needs_text.insert(tk.END, procurement_text)
                    else: # 如果缓存为空，确保显示默认示例
                        current_content = self.procurement_needs_text.get("1.0", tk.END).strip()
                        if not current_content or current_content.startswith("例如:"): # 避免重复插入示例
                            pass # 保持默认的示例
                        elif not current_content: # 如果getText返回的是空，插入示例
                             self.procurement_needs_text.delete("1.0", tk.END)
                             self.procurement_needs_text.insert(tk.END, "例如:\n土豆,70cm,100\n苹果,大,50\n白菜,,30")

        except FileNotFoundError:
            print(f"缓存文件 {cache_path} 未找到，将使用默认输入。")
            # 确保默认示例文本存在
            current_content = self.procurement_needs_text.get("1.0", tk.END).strip()
            if not current_content: # 如果文本框是空的，插入示例
                self.procurement_needs_text.insert(tk.END, "例如:\n土豆,70cm,100\n苹果,大,50\n白菜,,30")
        except json.JSONDecodeError:
            print(f"缓存文件 {cache_path} 格式错误，将使用默认输入。")
            os.remove(cache_path) # 删除损坏的缓存文件
            self.procurement_needs_text.delete("1.0", tk.END)
            self.procurement_needs_text.insert(tk.END, "例如:\n土豆,70cm,100\n苹果,大,50\n白菜,,30")
        except Exception as e:
            print(f"加载缓存失败: {e}")

    def _save_cached_procurement_list(self):
        cache_path = self._get_cache_file_path()
        procurement_text = self.procurement_needs_text.get("1.0", tk.END).strip()
        # 如果文本是默认的示例文字，可以考虑不缓存，或者缓存一个空字符串
        default_example = "例如:\n土豆,70cm,100\n苹果,大,50\n香蕉,小,20\n白菜,,30 (如无规格则第二项留空)"  # 与__init__中保持一致
        if procurement_text == default_example.strip():
            data_to_save = {"procurement_list": ""}  # 存空，下次加载时会用默认示例
        else:
            data_to_save = {"procurement_list": procurement_text}

        try:
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=4)
            print(f"采购清单已缓存到: {cache_path}")
        except Exception as e:
            print(f"保存缓存失败: {e}")
            messagebox.showwarning("缓存错误", f"无法保存采购清单到缓存文件。\n错误: {e}")

    def on_closing(self):
        """处理窗口关闭事件，保存缓存并退出"""
        self._save_cached_procurement_list()
        self.root.destroy()  # 关闭Tkinter窗口

    def clear_inputs_and_cache(self):  # 新方法，或修改原 clear_inputs
        """清空所有输入，并清除缓存文件"""
        # 先调用原来的清空逻辑 (如果它只清空UI)
        self.clear_inputs()  # 调用修改后的clear_inputs

        # 清除文本框并设置默认示例
        self.procurement_needs_text.delete("1.0", tk.END)
        self.procurement_needs_text.insert(tk.END, "例如:\n土豆,70cm,100\n苹果,大,50\n香蕉,小,20\n白菜,,30 (如无规格则第二项留空)")

        # 删除缓存文件
        cache_path = self._get_cache_file_path()
        try:
            if os.path.exists(cache_path):
                os.remove(cache_path)
                print(f"缓存文件 {cache_path} 已被清除。")
        except Exception as e:
            print(f"清除缓存文件失败: {e}")

        # 确保调用了 _save_cached_procurement_list 来保存当前的空/默认状态（如果需要）
        # 或者，如果希望“清空缓存”后下次打开是空白，则_save_cached_procurement_list中应处理空内容
        self._save_cached_procurement_list()  # 保存当前（可能是示例）状态到缓存

    def _add_supplier_row_ui(self, file_path="", supplier_name_val=""):
        # (基本不变)
        if len(self.supplier_entries) >= self.MAX_SUPPLIERS: return
        row_index = len(self.supplier_entries)
        row_frame = ttk.Frame(self.suppliers_dynamic_area)
        row_frame.pack(fill=tk.X, pady=2)
        self.supplier_frame_widgets[row_index] = row_frame
        name_var = tk.StringVar()
        default_name = supplier_name_val if supplier_name_val else f"供应商 {chr(ord('A') + row_index)}"
        if file_path:
            base = os.path.basename(file_path)
            potential_name = os.path.splitext(base)[0]
            name_var.set(potential_name if potential_name else default_name)
        else:
            name_var.set(default_name)
        name_label = ttk.Label(row_frame, text="名称:")
        name_label.grid(row=0, column=0, padx=(0, 2), pady=2, sticky=tk.W)
        name_entry = ttk.Entry(row_frame, textvariable=name_var, width=18)
        name_entry.grid(row=0, column=1, padx=(0, 5), pady=2, sticky=tk.W)
        file_label_text = "报价单:"
        file_label = ttk.Label(row_frame, text=file_label_text)
        file_label.grid(row=0, column=2, padx=(0, 2), pady=2, sticky=tk.W)
        path_var = tk.StringVar(value=file_path)
        file_entry = ttk.Entry(row_frame, textvariable=path_var, width=40, state='readonly')
        file_entry.grid(row=0, column=3, padx=(0, 5), pady=2, sticky=tk.EW)
        browse_button = ttk.Button(row_frame, text="选择文件", command=lambda idx=row_index: self.browse_file_for_supplier(idx))
        browse_button.grid(row=0, column=4, pady=2, sticky=tk.W)
        row_frame.grid_columnconfigure(3, weight=1)
        self.supplier_entries.append({
            'name_label': name_label, 'name_var': name_var, 'name_entry': name_entry,
            'file_label': file_label, 'file_entry': file_entry,
            'browse_button': browse_button, 'path_var': path_var,
            'row_frame': row_frame
        })
        self._update_add_remove_buttons_state()

    def add_supplier_input(self):
        if len(self.supplier_entries) < self.MAX_SUPPLIERS: self._add_supplier_row_ui()

    def remove_last_supplier_input(self):
        if len(self.supplier_entries) > self.MIN_SUPPLIERS:
            last_supplier_widgets = self.supplier_entries.pop()
            last_supplier_widgets['row_frame'].destroy()
            del self.supplier_frame_widgets[len(self.supplier_entries)]
            self._update_add_remove_buttons_state()
            if self.purchase_table: self.rebuild_treeview_columns()

    def _update_add_remove_buttons_state(self):
        num_suppliers = len(self.supplier_entries)
        self.add_supplier_button.config(state=tk.NORMAL if num_suppliers < self.MAX_SUPPLIERS else tk.DISABLED)
        self.remove_supplier_button.config(state=tk.NORMAL if num_suppliers > self.MIN_SUPPLIERS else tk.DISABLED)

    def browse_file_for_supplier(self, supplier_index):
        path_var = self.supplier_entries[supplier_index]['path_var']
        name_var = self.supplier_entries[supplier_index]['name_var']
        file_path = filedialog.askopenfilename(title=f"选择 {name_var.get()} 的报价单", filetypes=(("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")))
        if file_path:
            path_var.set(file_path)
            base = os.path.basename(file_path)
            potential_name = os.path.splitext(base)[0]
            potential_name = potential_name.replace("报价单", "").replace("报价", "").replace("价格表", "").strip()
            if not potential_name: potential_name = f"供应商 {chr(ord('A') + supplier_index)}"
            name_var.set(potential_name)

    def _setup_purchase_table(self, supplier_display_names_for_cols):
        # (基本不变, 包含比价参考列)
        if self.purchase_table:
            self.purchase_table.destroy()
            for widget in self.purchase_table_container.winfo_children(): widget.destroy()
        fixed_cols_before_comparison = ('采购数量', '单价', '金额')
        separator_col = ('比价过程参考',)
        supplier_quote_cols = tuple(f"{name}报价" for name in supplier_display_names_for_cols)
        tree_columns_data = fixed_cols_before_comparison + separator_col + supplier_quote_cols
        self.purchase_table = ttk.Treeview(self.purchase_table_container, columns=tree_columns_data, show='headings tree', height=15)
        self.purchase_table.heading("#0", text="供应商 / 产品名称 (规格)")
        self.purchase_table.column("#0", width=280, anchor=tk.W, stretch=tk.YES)
        for col_name in tree_columns_data:
            self.purchase_table.heading(col_name, text=col_name)
            if col_name == "比价过程参考":
                self.purchase_table.column(col_name, width=120, anchor=tk.CENTER, stretch=tk.NO)
            elif "报价" in col_name:
                self.purchase_table.column(col_name, width=90, anchor=tk.CENTER, stretch=tk.NO)
            elif col_name == "采购数量":
                self.purchase_table.column(col_name, width=70, anchor=tk.CENTER, stretch=tk.NO)
            elif col_name == "单价" or col_name == "金额":
                self.purchase_table.column(col_name, width=80, anchor=tk.E, stretch=tk.NO)
        vsb = ttk.Scrollbar(self.purchase_table_container, orient="vertical", command=self.purchase_table.yview)
        hsb = ttk.Scrollbar(self.purchase_table_container, orient="horizontal", command=self.purchase_table.xview)
        self.purchase_table.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.purchase_table.grid(row=0, column=0, sticky='nsew');
        vsb.grid(row=0, column=1, sticky='ns');
        hsb.grid(row=1, column=0, sticky='ew')
        self.purchase_table_container.grid_rowconfigure(0, weight=1);
        self.purchase_table_container.grid_columnconfigure(0, weight=1)
        self.purchase_table.tag_configure('supplier_header', font='-weight bold')

    def rebuild_treeview_columns(self):
        current_supplier_display_names = [s_entry['name_var'].get() for s_entry in self.supplier_entries if s_entry['path_var'].get()]
        if not current_supplier_display_names and self.supplier_entries:
            current_supplier_display_names = [s_entry['name_var'].get() for s_entry in self.supplier_entries]

        self._setup_purchase_table(current_supplier_display_names if current_supplier_display_names else [f"供应商 {chr(ord('A') + i)}" for i in range(len(self.supplier_entries))])

        if self.purchase_table:  # 确保表格已创建
            for item in self.purchase_table.get_children():
                self.purchase_table.delete(item)

        self.export_button.config(state=tk.DISABLED)
        self.current_purchase_df = pd.DataFrame()  # 清空数据

    def clear_inputs(self):
        for _ in range(len(self.supplier_entries) - self.INITIAL_SUPPLIERS):
            if len(self.supplier_entries) > self.INITIAL_SUPPLIERS:
                self.remove_last_supplier_input()
            elif len(self.supplier_entries) < self.INITIAL_SUPPLIERS:
                for _ in range(self.INITIAL_SUPPLIERS - len(self.supplier_entries)): self._add_supplier_row_ui()
        for i in range(min(len(self.supplier_entries), self.INITIAL_SUPPLIERS)):
            self.supplier_entries[i]['path_var'].set("")
            self.supplier_entries[i]['name_var'].set(f"供应商 {chr(ord('A') + i)}")

        self.product_name_col_var.set(self.DEFAULT_PRODUCT_NAME_COL)
        self.spec_col_var.set(self.DEFAULT_SPEC_COL)
        self.price_col_var.set(self.DEFAULT_PRICE_COL)

        # self.procurement_needs_text.delete("1.0", tk.END)
        # self.procurement_needs_text.insert(tk.END, "例如:\n土豆,70cm,100\n苹果,大,50\n香蕉,小,20\n白菜,,30 (如无规格则第二项留空)")

        if self.purchase_table:
            for item in self.purchase_table.get_children(): self.purchase_table.delete(item)

        # 移除对已删除控件的引用
        # self.supplier_totals_text.config(state=tk.NORMAL); self.supplier_totals_text.delete("1.0", tk.END); self.supplier_totals_text.config(state=tk.DISABLED)
        # self.notes_text.config(state=tk.NORMAL); self.notes_text.delete("1.0", tk.END); self.notes_text.config(state=tk.DISABLED)
        # self.total_procurement_cost_var.set("总采购额: 0.00 元")

        self.export_button.config(state=tk.DISABLED)
        self.current_purchase_df = pd.DataFrame()
        self._update_add_remove_buttons_state()
        self.rebuild_treeview_columns()  # 确保表格列与初始供应商状态一致

    def run_analysis(self):
        active_suppliers_info = []
        for i, s_entry in enumerate(self.supplier_entries):
            file_path = s_entry['path_var'].get()
            display_name = s_entry['name_var'].get().strip()
            if not display_name:
                display_name = f"供应商 {chr(ord('A') + i)}"
                s_entry['name_var'].set(display_name)
            if file_path:
                active_suppliers_info.append({'path': file_path, 'name': display_name})
        if not active_suppliers_info or len(active_suppliers_info) < self.MIN_SUPPLIERS:
            messagebox.showerror('输入错误', f'请至少为 {self.MIN_SUPPLIERS} 个供应商选择报价文件并确保它们有名称！')
            return

        current_supplier_display_names_for_cols = [s_info['name'] for s_info in active_suppliers_info]
        self._setup_purchase_table(current_supplier_display_names_for_cols)

        procurement_text_input = self.procurement_needs_text.get("1.0", tk.END)
        product_name_col = self.product_name_col_var.get().strip()
        spec_col = self.spec_col_var.get().strip()  # 如果为空字符串，表示用户不使用规格列
        price_col = self.price_col_var.get().strip()

        if not (product_name_col and price_col):
            messagebox.showerror('输入错误', '请输入Excel中的品名列名和价格列名！')
            return

        # 如果用户没有输入规格列名，spec_col会是空字符串
        # load_and_prepare_data 会将 None 或空字符串的 spec_col_name 视为无规格列

        procurement_needs_internal, _ = parse_procurement_input(procurement_text_input)  # procurement_needs_display 暂时不用
        if procurement_needs_internal is None: return

        loaded_dfs_dict = {}
        for s_info in active_suppliers_info:
            # 传递 spec_col (可能是空字符串) 给 load_and_prepare_data
            df = load_and_prepare_data(s_info['path'], s_info['name'], product_name_col, spec_col if spec_col else None, price_col)
            if df is None: return
            loaded_dfs_dict[s_info['name']] = df

        purchase_df, supplier_totals_dict, notes_list = generate_purchase_plan(
            loaded_dfs_dict,
            procurement_needs_internal,
            current_supplier_display_names_for_cols
        )
        self.current_purchase_df = purchase_df

        # 保存本次运行的汇总数据，以便导出Excel时使用
        self.last_run_supplier_totals_dict = supplier_totals_dict
        self.last_run_notes_list = notes_list
        # self.last_run_grand_total_cost 会在下面计算

        if self.purchase_table:
            for item in self.purchase_table.get_children(): self.purchase_table.delete(item)

        grand_total_cost = 0.0  # 重置grand_total_cost
        if not purchase_df.empty:
            for chosen_sup_name in current_supplier_display_names_for_cols:
                group_df = purchase_df[purchase_df['选择的供应商'] == chosen_sup_name]
                if not group_df.empty:
                    total_for_supplier = supplier_totals_dict.get(chosen_sup_name, 0.0)
                    grand_total_cost += total_for_supplier
                    parent_iid = self.purchase_table.insert("", tk.END,
                                                            text=f"{chosen_sup_name} (总计: {total_for_supplier:.2f} 元)",
                                                            open=True, tags=('supplier_header',))
                    for index, row in group_df.iterrows():
                        product_display_name_from_df = row['产品显示名称']
                        qty = row['采购数量']
                        unit_price_selected = row['单价']
                        item_total = row['金额']
                        fixed_values = (qty, f"{unit_price_selected:.2f}", f"{item_total:.2f}")
                        separator_value = ("---->",)
                        comparison_values = []
                        comparison_details_dict = row['比价详情']
                        for sup_compare_name in current_supplier_display_names_for_cols:
                            price = comparison_details_dict.get(sup_compare_name, "未报价")
                            formatted_price = f"{price:.2f}" if isinstance(price, (int, float)) else str(price)
                            comparison_values.append(formatted_price)
                        child_values = fixed_values + separator_value + tuple(comparison_values)
                        self.purchase_table.insert(parent_iid, tk.END, text=f"  └ {product_display_name_from_df}", values=child_values)
            self.export_button.config(state=tk.NORMAL)
        else:
            if not notes_list or all("没有加载到任何有效的供应商报价数据。" in note for note in notes_list) or all("所有供应商的报价数据均为空或无效。" in note for note in notes_list):
                if not any("以下产品在所有供应商报价中均未找到" in note for note in notes_list):
                    messagebox.showinfo("结果提示", "未能生成任何采购条目。\n请检查输入和供应商数据。")
            self.export_button.config(state=tk.DISABLED)

        self.last_run_grand_total_cost = grand_total_cost  # 保存计算出的总额

        # 移除对已删除UI控件的更新
        # self.total_procurement_cost_var.set(f"总采购额: {grand_total_cost:.2f} 元")
        # self.supplier_totals_text.config(state=tk.NORMAL); self.supplier_totals_text.delete("1.0", tk.END)
        # ...
        # self.notes_text.config(state=tk.NORMAL); self.notes_text.delete("1.0", tk.END)
        # ...

    def export_to_excel(self):
        if self.current_purchase_df.empty:
            messagebox.showerror("导出错误", "没有可导出的采购数据。")
            return
        try:
            today_date_str = datetime.datetime.now().strftime("%Y-%m-%d")
            default_filename = f"{today_date_str}_采购清单.xlsx"
            save_path = filedialog.asksaveasfilename(title='将采购单另存为 Excel 文件', defaultextension=".xlsx", initialfile=default_filename, filetypes=(("Excel 文件", "*.xlsx"), ("所有文件", "*.*")))
            if not save_path: return

            wb = Workbook()
            ws = wb.active
            ws.title = "采购明细(按供应商)"
            header_font = Font(bold=True, name='Arial', size=11)
            supplier_header_font = Font(bold=True, name='Arial', size=11, color="00008B")
            product_font = Font(name='Arial', size=10)
            center_alignment = Alignment(horizontal='center', vertical='center')
            right_alignment = Alignment(horizontal='right', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            border_bottom_thin = Border(bottom=Side(style='thin'))
            supplier_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

            current_supplier_display_names_for_export = [s_entry['name_var'].get() for s_entry in self.supplier_entries if s_entry['path_var'].get()]
            if not current_supplier_display_names_for_export:
                current_supplier_display_names_for_export = [s_entry['name_var'].get() for s_entry in self.supplier_entries]

            fixed_headers = ["供应商 / 产品", "品名", "规格", "采购数量", "单价", "金额"]
            separator_header = ["比价过程参考"]
            supplier_quote_headers = [f"{name}报价" for name in current_supplier_display_names_for_export]
            headers = fixed_headers + separator_header + supplier_quote_headers
            ws.append(headers)
            for col_num, header_title in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font;
                cell.alignment = center_alignment;
                cell.border = border_bottom_thin
                if header_title == "供应商 / 产品":
                    ws.column_dimensions[get_column_letter(col_num)].width = 30
                elif header_title == "品名":
                    ws.column_dimensions[get_column_letter(col_num)].width = 20
                elif header_title == "规格":
                    ws.column_dimensions[get_column_letter(col_num)].width = 15
                elif header_title == "比价过程参考":
                    ws.column_dimensions[get_column_letter(col_num)].width = 18
                elif "报价" in header_title:
                    ws.column_dimensions[get_column_letter(col_num)].width = 12
                else:
                    ws.column_dimensions[get_column_letter(col_num)].width = 10

            current_row = 2
            # grand_total_export = 0.0 # 使用 self.last_run_grand_total_cost

            if not self.current_purchase_df.empty:
                for chosen_sup_name in current_supplier_display_names_for_export:
                    group_df = self.current_purchase_df[self.current_purchase_df['选择的供应商'] == chosen_sup_name]
                    if not group_df.empty:
                        supplier_total_amount = group_df['金额'].sum()
                        # grand_total_export += supplier_total_amount # 不再在这里累加，直接用保存的总额
                        ws.cell(row=current_row, column=1, value=f"{chosen_sup_name} (总计: {supplier_total_amount:.2f} 元)").font = supplier_header_font
                        ws.cell(row=current_row, column=1).fill = supplier_fill
                        current_row += 1
                        for index, row_data in group_df.iterrows():
                            product_name_from_df = row_data['品名']
                            spec_from_df = row_data['规格'] if pd.notna(row_data['规格']) else ""
                            qty = row_data['采购数量']
                            unit_price = row_data['单价']
                            amount = row_data['金额']
                            comparison_details = row_data['比价详情']
                            excel_fixed_values = [f"  └ {row_data['产品显示名称']}", product_name_from_df, spec_from_df, qty, f"{unit_price:.2f}", f"{amount:.2f}"]
                            excel_separator_value = ["---->"]
                            excel_comparison_values = []
                            for sup_compare_name in current_supplier_display_names_for_export:
                                price = comparison_details.get(sup_compare_name, "未报价")
                                excel_comparison_values.append(f"{price:.2f}" if isinstance(price, (int, float)) else str(price))
                            excel_row_values = excel_fixed_values + excel_separator_value + excel_comparison_values
                            ws.append(excel_row_values)
                            for col_idx, value_val in enumerate(excel_row_values, 1):
                                cell = ws.cell(row=current_row, column=col_idx)
                                cell.font = product_font
                                current_header = headers[col_idx - 1]
                                if current_header == "供应商 / 产品":
                                    cell.alignment = left_alignment
                                elif current_header in ["品名", "规格", "比价过程参考"] or isinstance(value_val, str) and value_val == "未报价":
                                    cell.alignment = center_alignment
                                elif isinstance(value_val, (int, float)) or (isinstance(value_val, str) and value_val.replace('.', '', 1).replace('-', '', 1).isdigit()):
                                    cell.alignment = right_alignment
                                else:
                                    cell.alignment = center_alignment
                            current_row += 1
                        current_row += 1

            # 写入总采购额和备注 (如果需要)
            ws.cell(row=current_row, column=len(headers) - 1, value="总采购额:").font = header_font
            ws.cell(row=current_row, column=len(headers) - 1).alignment = right_alignment
            ws.cell(row=current_row, column=len(headers), value=f"{self.last_run_grand_total_cost:.2f}").font = header_font  # 使用保存的总额
            ws.cell(row=current_row, column=len(headers)).alignment = right_alignment
            current_row += 2

            notes_to_export = self.last_run_notes_list
            if notes_to_export and (len(notes_to_export) > 1 or (len(notes_to_export) == 1 and notes_to_export[0] != "无特殊备注信息.")):
                ws.cell(row=current_row, column=1, value="备注信息:").font = header_font
                current_row += 1
                for note_line in notes_to_export:
                    ws.cell(row=current_row, column=1, value=note_line).font = product_font
                    current_row += 1

            wb.save(save_path)
            messagebox.showinfo("导出成功", "采购单已成功导出!")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出Excel失败: {e}")


if __name__ == '__main__':
    root = tk.Tk()
    app = ProcurementApp(root)
    root.mainloop()